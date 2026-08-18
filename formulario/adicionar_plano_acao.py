# -*- coding: utf-8 -*-
"""
Adiciona as colunas de PLANO DE AÇÃO ao formulário de maturidade.

Gera uma cópia do formulário F_O_025 com 4 novas colunas nas abas
Avaliação, Indicadores e Treinamento:
    Plano de Ação | Responsável | Prazo | Status da Ação

O arquivo gerado é a fonte de dados do Power BI (Power Query lê direto dele).
"""
import glob
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ORIGEM_GLOB = r"C:\Users\User\Downloads\F_O_025_GOCO*"
DESTINO = Path(__file__).resolve().parent / "F_O_025_Formulario_Maturidade_Planos.xlsx"

ABAS = ["Avaliação", "Indicadores", "Treinamento"]
COLUNAS_NOVAS = ["Plano de Ação", "Responsável", "Prazo", "Status da Ação"]
STATUS_VALIDOS = "Aberto,Em andamento,Concluído"


def main():
    arquivos = glob.glob(ORIGEM_GLOB)
    if not arquivos:
        raise SystemExit("Formulário de origem não encontrado em Downloads.")
    origem = arquivos[0]

    shutil.copy2(origem, DESTINO)

    wb = load_workbook(DESTINO)

    for aba in ABAS:
        ws = wb[aba]
        cabecalho = [c.value for c in ws[1]]
        if "Plano de Ação" in cabecalho:
            continue

        indice_obs = cabecalho.index("Observação") + 1
        for i, nome in enumerate(COLUNAS_NOVAS, start=1):
            col = indice_obs + i
            ws.cell(row=1, column=col, value=nome)

        col_status = indice_obs + len(COLUNAS_NOVAS)
        letra_status = get_column_letter(col_status)
        dv = DataValidation(
            type="list",
            formula1=f'"{STATUS_VALIDOS}"',
            allow_blank=True,
            showDropDown=False,
        )
        ws.add_data_validation(dv)
        dv.add(f"{letra_status}2:{letra_status}{ws.max_row}")

    wb.save(DESTINO)
    print("OK ->", DESTINO)


if __name__ == "__main__":
    main()